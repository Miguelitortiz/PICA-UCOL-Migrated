#!/usr/bin/env python3
"""
Script para actualizar tutores y aulas en class_groups desde el Excel
'info para horarios.xlsx'

Hoja 1 (tutores_aulas_matutino): tutor + aula fija por grupo matutino
Hoja 2 (aulas vespertino): aulas por día para grupos vespertinos

Acciones:
1. Agrega columnas classroom y classrooms_by_day a class_groups si no existen
2. Hace match de nombres tutor (Excel) con professors en la DB
3. Actualiza tutor_id, classroom y classrooms_by_day en class_groups
"""

import psycopg2
import openpyxl
import json
import unicodedata
import re
import sys

DB_URL = 'postgres://admin:admin_pass@localhost:5432/pica_db'

# Mapping de abreviaturas de carrera del Excel → career_id en la DB
# IME=371, ICI=349, ISET=99, IMT=418
CAREER_MAP = {
    'IME': 371,
    'ICI': 349,
    'ISET': 99,
    'IMT': 418,
}


def normalize_name(name: str) -> str:
    """Normaliza un nombre para comparación: minúsculas, sin acentos, sin espacios extra."""
    if not name:
        return ''
    name = name.strip()
    # Remove accents
    nfkd = unicodedata.normalize('NFKD', name)
    ascii_name = ''.join(c for c in nfkd if not unicodedata.combining(c))
    # Lowercase and remove non-alpha chars
    ascii_name = re.sub(r'[^a-z ]', '', ascii_name.lower())
    # Collapse spaces
    return ' '.join(ascii_name.split())


def name_tokens(name: str) -> set:
    """Retorna tokens de un nombre normalizado."""
    return set(normalize_name(name).split())


def best_professor_match(excel_name: str, professors: list[tuple]) -> tuple | None:
    """
    Intenta hacer match entre un nombre del Excel y la lista de profesores de la DB.
    La DB tiene formato 'Apellido Nombre', el Excel tiene 'Nombre Apellido'.
    Retorna (id, full_name) del mejor match o None.
    """
    excel_tokens = name_tokens(excel_name)
    
    best_match = None
    best_score = 0
    
    for prof_id, prof_name in professors:
        prof_tokens = name_tokens(prof_name)
        # Score: número de tokens en común / tokens únicos (Jaccard similarity)
        intersection = len(excel_tokens & prof_tokens)
        union = len(excel_tokens | prof_tokens)
        score = intersection / union if union > 0 else 0
        
        if score > best_score:
            best_score = score
            best_match = (prof_id, prof_name)
    
    # Solo retornar si hay al menos 2 tokens en común (evitar falsos positivos)
    if best_match and best_score >= 0.4:
        return best_match
    return None


def parse_matutino_sheet(ws):
    """
    Parsea la hoja de matutino.
    Retorna lista de dicts: {grupo, aula, tutor_name, carrera_abbr}
    """
    groups = []
    rows = list(ws.iter_rows(values_only=True))
    
    # Estructura: fila 0=career headers, fila 1=column headers, fila 2+=datos
    # Columnas (0-indexed): 
    #   IME:  grupo=2, aula=3, tutor=4
    #   ICI:  grupo=6, aula=7, tutor=8
    #   ISET: grupo=10, aula=11, tutor=12
    #   IMT:  grupo=14, aula=15, tutor=16
    
    COLS = [
        ('IME',  2,  3,  4),
        ('ICI',  6,  7,  8),
        ('ISET', 10, 11, 12),
        ('IMT',  14, 15, 16),
    ]
    
    for row in rows[2:]:  # Skip header rows
        if not any(cell is not None for cell in row):
            continue
        
        for carrera, gc, ac, tc in COLS:
            grupo_val = row[gc] if gc < len(row) else None
            if not grupo_val:
                continue
            grupo_val = str(grupo_val).strip()
            # Skip header rows that accidentally got parsed
            if grupo_val in ('Grupo', carrera):
                continue
            
            aula = str(row[ac]).strip() if ac < len(row) and row[ac] else None
            if aula == 'None':
                aula = None
            tutor = str(row[tc]).strip() if tc < len(row) and row[tc] else None
            if tutor == 'None':
                tutor = None
            
            groups.append({
                'grupo': grupo_val,
                'aula': aula,
                'tutor_name': tutor,
                'carrera_abbr': carrera,
            })
    
    return groups


def parse_vespertino_sheet(ws):
    """
    Parsea la hoja vespertina.
    Retorna lista de dicts: {grupo, aulas_por_dia: {Lunes:..., Martes:..., ...}, carrera_abbr}
    """
    groups = []
    rows = list(ws.iter_rows(values_only=True))
    
    # Estructura:
    # fila 0: vacía
    # fila 1: "Ciclo 2026-2"
    # fila 2: "Aulas"
    # fila 3: "Turno Vespertino"
    # fila 4: IME (col 0), ISET (col 7), IMT (col 14)
    # fila 5: vacía
    # fila 6: Grupo, Día/Aula headers
    # fila 7: None, Lunes, Martes, Miércoles, Jueves, Viernes, None, None, Lunes...
    # fila 8: vacía
    # fila 9+: datos
    
    # Bloques de columnas (grupo_col, lunes_col, martes_col, mierc_col, juev_col, vier_col)
    BLOCKS = [
        ('IME',  0, 1, 2, 3, 4, 5),
        ('ISET', 7, 8, 9, 10, 11, 12),
        ('IMT',  14, 15, 16, 17, 18, 19),
    ]
    
    DIAS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
    
    # Find data start row (after the day names row)
    data_start = None
    for i, row in enumerate(rows):
        if row and row[1] == 'Lunes':
            data_start = i + 1
            break
    
    if data_start is None:
        print("WARN: No se encontró la fila de días en la hoja vespertino")
        return groups
    
    # Skip empty rows after day names
    for row in rows[data_start:]:
        if not any(cell is not None for cell in row):
            continue
        
        for carrera, gc, lc, mc, wc, jc, vc in BLOCKS:
            grupo_val = row[gc] if gc < len(row) else None
            if not grupo_val:
                continue
            grupo_val = str(grupo_val).strip()
            if grupo_val in ('Grupo', carrera, 'None'):
                continue
            
            day_cols = [lc, mc, wc, jc, vc]
            aulas_por_dia = {}
            for dia, col in zip(DIAS, day_cols):
                val = row[col] if col < len(row) else None
                if val and str(val).strip() not in ('None', ''):
                    aulas_por_dia[dia] = str(val).strip()
                else:
                    aulas_por_dia[dia] = None
            
            groups.append({
                'grupo': grupo_val,
                'aulas_por_dia': aulas_por_dia,
                'carrera_abbr': carrera,
            })
    
    return groups


def group_key_from_db(name: str) -> str:
    """
    Normaliza el nombre del grupo de la DB para comparar con el del Excel.
    Ej: '1°A' → '1A', '7°J' → '7J', '7ªB' → '7B'
    """
    name = name.strip()
    # Remove degree/superscript symbols
    name = re.sub(r'[°ª]', '', name)
    # Remove spaces
    name = re.sub(r'\s+', '', name)
    return name.upper()


def group_key_from_excel(name: str) -> str:
    """
    Normaliza el nombre del grupo del Excel para comparar.
    Ej: '1A' → '1A', '7 J' → '7J'
    """
    name = str(name).strip()
    # Remove spaces
    name = re.sub(r'\s+', '', name)
    return name.upper()


def main():
    print("=== Actualizando tutores y aulas de grupos ===\n")
    
    # Load Excel
    wb = openpyxl.load_workbook('info para horarios.xlsx')
    ws_matutino = wb['tutores_aulas_matutino']
    ws_vespertino = wb['aulas vespertino']
    
    matutino_groups = parse_matutino_sheet(ws_matutino)
    vespertino_groups = parse_vespertino_sheet(ws_vespertino)
    
    print(f"Grupos matutinos en Excel: {len(matutino_groups)}")
    print(f"Grupos vespertinos en Excel: {len(vespertino_groups)}")
    
    # Connect to DB
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()
    
    # Step 1: Ensure columns exist in class_groups
    print("\n--- Verificando columnas en class_groups ---")
    
    cur.execute("""
        SELECT column_name FROM information_schema.columns 
        WHERE table_name='class_groups'
    """)
    existing_cols = {r[0] for r in cur.fetchall()}
    
    if 'classroom' not in existing_cols:
        print("  Agregando columna 'classroom'...")
        cur.execute("ALTER TABLE class_groups ADD COLUMN classroom VARCHAR(50)")
    else:
        print("  Columna 'classroom' ya existe")
    
    if 'classrooms_by_day' not in existing_cols:
        print("  Agregando columna 'classrooms_by_day' (JSONB)...")
        cur.execute("ALTER TABLE class_groups ADD COLUMN classrooms_by_day JSONB")
    else:
        print("  Columna 'classrooms_by_day' ya existe")
    
    conn.commit()
    
    # Step 2: Load all DB groups and professors
    cur.execute("""
        SELECT id, name, slug, shift, career_id 
        FROM class_groups 
        ORDER BY name
    """)
    db_groups = cur.fetchall()
    
    cur.execute("SELECT id, full_name FROM professors ORDER BY full_name")
    db_professors = cur.fetchall()
    
    # Build lookup: (career_id, group_key) → db_group_id
    group_lookup = {}
    for gid, gname, gslug, gshift, gcareer_id in db_groups:
        key = (gcareer_id, group_key_from_db(gname))
        group_lookup[key] = (gid, gname, gshift)
    
    print(f"\n--- Procesando {len(matutino_groups)} grupos MATUTINOS ---")
    
    updated = 0
    skipped = 0
    no_tutor = 0
    
    for g in matutino_groups:
        career_id = CAREER_MAP.get(g['carrera_abbr'])
        if not career_id:
            print(f"  WARN: Carrera desconocida '{g['carrera_abbr']}'")
            continue
        
        gkey = group_key_from_excel(g['grupo'])
        lookup_key = (career_id, gkey)
        
        db_entry = group_lookup.get(lookup_key)
        if not db_entry:
            print(f"  WARN: Grupo no encontrado en DB: {g['carrera_abbr']} {g['grupo']} → clave {gkey}")
            skipped += 1
            continue
        
        db_id, db_name, db_shift = db_entry
        
        # Find tutor in DB
        tutor_id = None
        tutor_matched = None
        if g['tutor_name']:
            match = best_professor_match(g['tutor_name'], db_professors)
            if match:
                tutor_id, tutor_matched = match
            else:
                print(f"  WARN: No se encontró tutor para '{g['tutor_name']}' (grupo {db_name})")
                no_tutor += 1
        
        # Update group
        cur.execute("""
            UPDATE class_groups 
            SET tutor_id = %s, classroom = %s
            WHERE id = %s
        """, (tutor_id, g['aula'], db_id))
        
        status = f"tutor={tutor_matched or 'N/A'}, aula={g['aula'] or 'N/A'}"
        print(f"  ✓ {g['carrera_abbr']} {g['grupo']} ({db_name}): {status}")
        updated += 1
    
    conn.commit()
    print(f"\n  Matutinos: {updated} actualizados, {skipped} no encontrados, {no_tutor} sin tutor match")
    
    print(f"\n--- Procesando {len(vespertino_groups)} grupos VESPERTINOS ---")
    
    v_updated = 0
    v_skipped = 0
    
    for g in vespertino_groups:
        career_id = CAREER_MAP.get(g['carrera_abbr'])
        if not career_id:
            print(f"  WARN: Carrera desconocida '{g['carrera_abbr']}'")
            continue
        
        gkey = group_key_from_excel(g['grupo'])
        lookup_key = (career_id, gkey)
        
        db_entry = group_lookup.get(lookup_key)
        if not db_entry:
            print(f"  WARN: Grupo no encontrado en DB: {g['carrera_abbr']} {g['grupo']} → clave {gkey}")
            v_skipped += 1
            continue
        
        db_id, db_name, db_shift = db_entry
        
        # Filter out None values from aulas_por_dia
        aulas_clean = {k: v for k, v in g['aulas_por_dia'].items() if v}
        
        cur.execute("""
            UPDATE class_groups 
            SET classrooms_by_day = %s
            WHERE id = %s
        """, (json.dumps(aulas_clean, ensure_ascii=False), db_id))
        
        aulas_str = ', '.join(f"{d[:2]}: {a}" for d, a in aulas_clean.items())
        print(f"  ✓ {g['carrera_abbr']} {g['grupo']} ({db_name}): {aulas_str}")
        v_updated += 1
    
    conn.commit()
    print(f"\n  Vespertinos: {v_updated} actualizados, {v_skipped} no encontrados")
    
    # Final summary
    print("\n=== Resumen final ===")
    cur.execute("""
        SELECT name, shift, classroom, classrooms_by_day,
               (SELECT full_name FROM professors WHERE id = tutor_id) as tutor_name
        FROM class_groups
        ORDER BY shift DESC, name
    """)
    rows = cur.fetchall()
    print(f"Total grupos: {len(rows)}")
    for r in rows:
        name, shift, classroom, cbd, tutor = r
        print(f"  {name} ({shift}): tutor={tutor or '—'}, aula={classroom or cbd or '—'}")
    
    conn.close()
    print("\n✅ Completado.")


if __name__ == '__main__':
    main()
